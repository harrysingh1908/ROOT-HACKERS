import openpyxl
import openai

# Set up OpenAI API key
openai.api_key = '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'  # Replace with your actual OpenAI API key

def simple_chatbot(user_input):
    # ... (your existing code)
    responses = {
        "hello": "Hi there! How can I help you?",
        "how are you": "I'm just a computer program, but I'm doing well. Thanks for asking!",
        "bye": "Goodbye! Have a great day.",
        "what are dark patterns": "Dark patterns are deceptive user interface design tricks that manipulate users into taking actions they might not otherwise perform.",
        "how can we find dark patterns": "To find dark patterns, you can look for misleading information, confusing language, or interfaces that push you towards certain actions. Also, check for hidden checkboxes or buttons.",
        "what are some common types of dark patterns": "Common types of dark patterns include misdirection, hidden costs, forced continuity, and bait and switch.",
        "how do dark patterns affect user experience": "Dark patterns can negatively impact user experience by tricking users into unwanted actions, leading to frustration and mistrust.",
        "how can users protect themselves from dark patterns": "Users can protect themselves by staying vigilant, reading carefully, and being aware of common dark pattern tactics.",
        "default": "I'm not sure how to respond to that."
    }

    # Additional predefined questions related to dark patterns
    additional_dark_patterns_questions = {
        "give examples of misdirection in dark patterns": "Misdirection in dark patterns involves guiding users to take unintended actions. For example, a deceptive button might lead users to subscribe instead of canceling.",
        "how to identify hidden costs in interfaces": "Hidden costs in interfaces can be identified by carefully reviewing terms and conditions, checking for additional fees during the checkout process, and being wary of unclear pricing.",
        "explain forced continuity in dark patterns": "Forced continuity in dark patterns occurs when users are automatically enrolled in a subscription or service without clear consent. This can lead to unexpected charges.",
        "what is bait and switch in dark patterns": "Bait and switch in dark patterns involves enticing users with a desirable offer and then replacing it with a less desirable one. It's a deceptive tactic to lure users into making unintended choices."
    }

    # Convert user input to lowercase for case-insensitive matching
    user_input_lower = user_input.lower()

    # Check if the user's input is in the predefined responses
    response = responses.get(user_input_lower, None)

    if response is None:
        # Check if the user's input is in the additional predefined questions
        response = additional_dark_patterns_questions.get(user_input_lower, None)

    if response is None:
        # If the question is not in the predefined responses or additional questions, use ChatGPT to generate a response
        response = chatgpt_response(user_input)

    return response

def chatgpt_response(user_input):
    # Use ChatGPT to generate a response for the unknown question
    prompt = f"User: {user_input}\nChatbot:"
    response = openai.Completion.create(
        engine="text-davinci-002",
        prompt=prompt,
        max_tokens=150
    )
    return response['choices'][0]['text'].strip()

# Main loop for interacting with the chatbot
while True:
    user_input = input("You: ")
    if user_input.lower() == 'exit':
        rating = input("What was your experience with this website? Give a rating (1-10): ")

        # Load existing workbook if it exists
        try:
            wb = openpyxl.load_workbook('user_ratings.xlsx')
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            wb = openpyxl.Workbook()

        # Select the active sheet or create a new one
        sheet = wb.active if 'Sheet' in wb.sheetnames else wb.create_sheet("Sheet")

        # Write the header if the file is newly created
        if not sheet.iter_cols(min_row=1, max_row=1, max_col=1):
            sheet.append(["User Rating"])

        # Find the next available column and append the rating
        next_col = sheet.max_column + 1
        sheet.cell(row=1, column=next_col, value=f"Rating {next_col - 1}")
        sheet.cell(row=2, column=next_col, value=int(rating))  # Convert rating to integer

        # Calculate average rating
        average_rating = sum(sheet.cell(row=2, column=col).value for col in range(2, next_col)) / (next_col - 1)

        # Add column for average rating
        sheet.cell(row=1, column=next_col + 1, value="Average Rating")
        sheet.cell(row=2, column=next_col + 1, value=average_rating)

        # Save the workbook
        wb.save('user_ratings.xlsx')

        print(f"User Rating: {rating}")
        print(f"Average Rating: {average_rating}")
        print("Chatbot: Goodbye!")
        break

    bot_response = simple_chatbot(user_input)
    print("Chatbot:", bot_response)
